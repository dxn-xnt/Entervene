import argparse
import csv
import json
import re
import sys
import uuid
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

RAW_DIR = Path(__file__).parent.parent.parent / "data" / "raw_eclass_records"
OUTPUT_DIR = Path(__file__).parent.parent.parent / "data" / "datasets" / "new_extracted_pack"


def _generate_synthetic_id(name: str) -> str:
    """Generate a deterministic UUID based strictly on student name to link across years/sections."""
    clean_name = re.sub(r'[^a-z0-9]', '', name.lower())
    namespace = uuid.NAMESPACE_OID
    return str(uuid.uuid5(namespace, clean_name))


def extract_eclass_records():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    students = {}  # synthetic_id -> info
    period_grades = []
    
    for xlsx_path in sorted(RAW_DIR.glob("*.xlsx")):
        print(f"Processing {xlsx_path.name}...")
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        except Exception as e:
            print(f"  ERROR opening {xlsx_path.name}: {e}")
            continue
            
        # Determine grade, section, subject, school_year from filename
        # Default fallbacks
        filename_upper = xlsx_path.name.upper()
        grade_level = 10
        if " 7 " in filename_upper or "-7 " in filename_upper or "GRADE-7" in filename_upper: grade_level = 7
        elif " 8 " in filename_upper or "-8 " in filename_upper or "GRADE-8" in filename_upper: grade_level = 8
        elif " 9 " in filename_upper or "-9 " in filename_upper or "GRADE-9" in filename_upper or " 9-" in filename_upper: grade_level = 9
        elif "10" in filename_upper: grade_level = 10
        elif "11" in filename_upper: grade_level = 11
        elif "12" in filename_upper: grade_level = 12
        
        section = "UNKNOWN"
        if "EINSTEIN" in filename_upper: section = "EINSTEIN"
        elif "SOCRATES" in filename_upper: section = "SOCRATES"
        elif "ARCHIMEDES" in filename_upper: section = "ARCHIMEDES"
        elif "COPERNICUS" in filename_upper: section = "COPERNICUS"
        elif "NEWTON" in filename_upper: section = "NEWTON"
        elif "PLATO" in filename_upper: section = "PLATO"
        elif "ARISTOTLE" in filename_upper: section = "ARISTOTLE"
        elif "GALILEO" in filename_upper: section = "GALILEO"
        elif "CAMPOS" in filename_upper: section = "CAMPOS"
        elif "ZARA" in filename_upper: section = "ZARA"
        elif "DELMUNDO" in filename_upper: section = "DELMUNDO"
        
        subject = "UNKNOWN"
        if "PHYSICS" in filename_upper: subject = "ADVANCED_PHYSICS"
        elif "ELECTRONICS" in filename_upper: subject = "ELECTRONICS"
        elif "CON CHEM" in filename_upper: subject = "SCIENCE"
        elif "SCIENCE" in filename_upper: subject = "SCIENCE"
        elif "MATH" in filename_upper: subject = "MATHEMATICS"
        elif "CREATIVE TECH" in filename_upper: subject = "CREATIVE_TECHNOLOGY"
        elif "VALUES" in filename_upper: subject = "VALUES_EDUCATION"
        elif "ICT" in filename_upper: subject = "ICT"
        elif "MAPEH" in filename_upper: subject = "MAPEH"
        elif "ORAL COM" in filename_upper: subject = "ORAL_COMMUNICATION"
        elif "PERDEV" in filename_upper: subject = "PERSONAL_DEVELOPMENT"
        elif "PRECALCULUS" in filename_upper: subject = "PRE_CALCULUS"
        elif "ENGLISH" in filename_upper: subject = "ENGLISH"
        
        # Exact school year mapping based on user input
        school_year = "UNKNOWN"
        filename_exact = xlsx_path.name
        
        sy_2022_2023 = [
            "GRADE-10-mapeh-SOCRATES-FIRST-QUARTER-GRADES.xlsx",
            "Mapeh 10-Socrates-2nd-quarter-Grades-2022-23.xlsx"
        ]
        sy_2023_2024 = [
            "ADV. PHYSICS 10 - EINSTEIN.xlsx",
            "ADV. PHYSICS 10 - SOCRATES.xlsx",
            "CON CHEM 9- Archimedes.xlsx",
            "CON CHEM 9- Copernicus.xlsx",
            "ICT 9 - ARCHIMEDES.xlsx",
            "MATH 9-ARCHIMEDES.xlsx",
            "SCIENCE COPERICUS.xlsx",
            "Science 9 -4th Quarter Archimedes.xlsx",
            "Science 9 -4th Quarter Copernicus.xlsx",
            "MATH 10 - EINSTEIN (2023 - 2024).xlsx",
            "MATH 10 - SOCRATES (2023 - 2024).xlsx",
            "PRECALCULUS  11-ZARA - 1st.xlsx",
            "PRECALCULUS 11-CAMPOS - 1st.xlsx"
        ]
        sy_2024_2025 = [
            "MATH 10 -EINSTEIN 24-25.xlsx",
            "ORAL COM 11-CAMPOS .2024-2025.xlsx"
        ]
        sy_2025_2026 = [
            "CLASSRECORD. 7 - ARISTOTLE.xlsx",
            "CLASSRECORD. 7 -GALILEO.xlsx",
            "GRADE-10 EINSTEIN (ELECTRONICS).xlsx",
            "GRADE-10 SOCRATES (ELECTRONICS).xlsx",
            "GRADE-8 NEWTON (CREATIVE TECH).xlsx",
            "GRADE-8 NEWTON (VALUES EDUCATION).xlsx",
            "GRADE-8 PLATO (CREATIVE TECH).xlsx",
            "GRADE-8 PLATO (VALUES EDUCATION).xlsx",
            "GRADE-9 ARCHIMEDES (CREATIVE TECH).xlsx",
            "GRADE-9 COPERNICUS (CREATIVE TECH).xlsx",
            "PERDEV 12 - DELMUNDO.xlsx"
        ]
        
        if filename_exact in sy_2022_2023: school_year = "2022-2023"
        elif filename_exact in sy_2023_2024: school_year = "2023-2024"
        elif filename_exact in sy_2024_2025: school_year = "2024-2025"
        elif filename_exact in sy_2025_2026: school_year = "2025-2026"

        
        for sheet_name in wb.sheetnames:
            # We are looking for quarter/term sheets (Q1, Q2, Q3, Q4, 1ST, 2ND, etc.)
            sheet_upper = sheet_name.upper()
            period_seq = None
            if "Q1" in sheet_upper or "1ST" in sheet_upper: period_seq = 1
            elif "Q2" in sheet_upper or "2ND" in sheet_upper: period_seq = 2
            elif "Q3" in sheet_upper or "3RD" in sheet_upper: period_seq = 3
            elif "Q4" in sheet_upper or "4TH" in sheet_upper: period_seq = 4
            
            if period_seq is None:
                continue
                
            # SPECIAL RULE: Science 9 and Advanced Physics 10 teachers switched after Q1, so we don't have Q2-Q4
            if period_seq > 1:
                if subject == "SCIENCE" and grade_level == 9:
                    continue
                if subject == "ADVANCED_PHYSICS" and grade_level == 10:
                    continue
                
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            
            # Find the header row (has 'PS' multiple times)
            header_row_idx = -1
            ps_indices = []
            quarterly_grade_idx = -1
            initial_grade_idx = -1
            
            for r_idx, row in enumerate(all_rows[:15]):
                ps_found = [c_idx for c_idx, cell in enumerate(row) if str(cell).strip().upper() == "PS"]
                if len(ps_found) >= 2:
                    header_row_idx = r_idx
                    ps_indices = ps_found
                    # Usually "Quarterly Grade" or similar is at the end
                    for c_idx, cell in enumerate(row):
                        cell_str = str(cell).strip().upper()
                        if "QUARTERLY" in cell_str or "GRADE" in cell_str:
                            quarterly_grade_idx = c_idx
                        elif "INITIAL" in cell_str:
                            initial_grade_idx = c_idx
                    
                    if quarterly_grade_idx == -1:
                        # Fallback: look in the row above or below
                        for off in [-1, 1, -2, 2]:
                            if 0 <= r_idx + off < len(all_rows):
                                for c_idx, cell in enumerate(all_rows[r_idx + off]):
                                    cell_str = str(cell).strip().upper()
                                    if "QUARTERLY" in cell_str and "GRADE" in cell_str:
                                        quarterly_grade_idx = c_idx
                                    elif "INITIAL" in cell_str and "GRADE" in cell_str:
                                        initial_grade_idx = c_idx
                    break
            
            if header_row_idx == -1:
                continue
                
            # Fallback if Quarterly grade idx not found
            if quarterly_grade_idx == -1 and ps_indices:
                # It's usually the last column after the last PS
                quarterly_grade_idx = ps_indices[-1] + 3
                
            # Find MALE / FEMALE start rows
            student_rows = []
            in_students = False
            for r_idx in range(header_row_idx + 1, min(100, len(all_rows))):
                row = all_rows[r_idx]
                first_cell = str(row[0]).strip().upper() if row[0] is not None else ""
                second_cell = str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else ""
                
                if first_cell == "MALE" or second_cell == "MALE" or first_cell == "FEMALE" or second_cell == "FEMALE":
                    in_students = True
                    continue
                
                if first_cell == "" and second_cell == "":
                    continue # Skip empty rows
                
                # If we see a number in first column and string in second, it's a student
                if in_students:
                    name_cell = None
                    if str(row[0]).isdigit() and len(str(row[1]).strip()) > 3:
                        name_cell = str(row[1]).strip()
                    elif str(row[1]).isdigit() and len(row) > 2 and len(str(row[2]).strip()) > 3:
                        name_cell = str(row[2]).strip()
                    elif first_cell == "" and len(second_cell) > 3 and "," in second_cell:
                        name_cell = second_cell
                        
                    if name_cell and "," in name_cell:
                        student_rows.append((name_cell, row))
            
            for name, row in student_rows:
                synth_id = _generate_synthetic_id(name)
                if synth_id not in students:
                    students[synth_id] = {
                        "student_id": synth_id,
                        "name": name,
                        "grade_level": grade_level,
                        "section": section,
                        "school_year": school_year,
                    }
                
                # Extract grades
                ww_ps = pt_ps = qa_ps = q_grade = None
                
                try:
                    if len(ps_indices) >= 1 and ps_indices[0] < len(row):
                        ww_ps = float(row[ps_indices[0]])
                    if len(ps_indices) >= 2 and ps_indices[1] < len(row):
                        pt_ps = float(row[ps_indices[1]])
                    if len(ps_indices) >= 3 and ps_indices[2] < len(row):
                        qa_ps = float(row[ps_indices[2]])
                except (ValueError, TypeError):
                    pass
                    
                try:
                    if quarterly_grade_idx < len(row) and row[quarterly_grade_idx] is not None:
                        q_grade = float(row[quarterly_grade_idx])
                except (ValueError, TypeError):
                    pass
                
                # SPECIAL RULE: Skip if both written work and performance task scores are completely missing
                # This prevents extracting fake "60" transmuted grades from empty excel rows
                if ww_ps is None and pt_ps is None:
                    continue
                
                if q_grade is not None:
                    period_grades.append({
                        "student_id": synth_id,
                        "subject": subject,
                        "school_year": school_year,
                        "grade_level": grade_level,
                        "section": section,
                        "period_sequence": period_seq,
                        "written_work_percent": ww_ps if ww_ps is not None else "",
                        "performance_task_percent": pt_ps if pt_ps is not None else "",
                        "quarterly_assessment_percent": qa_ps if qa_ps is not None else "",
                        "source_period_grade": q_grade
                    })
        
        wb.close()
    
    # Save CSVs
    print(f"\nExtracted {len(students)} unique students and {len(period_grades)} period grades.")
    
    with open(OUTPUT_DIR / "students.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["student_id", "name", "grade_level", "section", "school_year"])
        writer.writeheader()
        for s in students.values():
            writer.writerow(s)
            
    with open(OUTPUT_DIR / "student_period_grades.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["student_id", "subject", "school_year", "grade_level", "section", "period_sequence", "written_work_percent", "performance_task_percent", "quarterly_assessment_percent", "source_period_grade"])
        writer.writeheader()
        for pg in period_grades:
            writer.writerow(pg)

    print(f"Data saved to {OUTPUT_DIR}")

if __name__ == "__main__":
    extract_eclass_records()
